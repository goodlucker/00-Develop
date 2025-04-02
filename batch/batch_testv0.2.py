import os
import subprocess
from openpyxl import load_workbook

# Set SAS paths and parameters
sas_path = r"C:\Program Files\SASHome\SASFoundation\9.4\sas.exe"
sas_user_path = r"C:\Program Files\SASHome\SASFoundation\9.4"
sas_ver = r"C:\Program Files\SASHome\SASFoundation\9.4\nls\u8\sasv9.cfg"

# Prompt user to select tracker Excel file
track_path = input("Please select the tracker Excel file: ")

# Load the tracker workbook
wb = load_workbook(filename=track_path)
sheet = wb.active

# Get batch type from user input
comb_type = int(input("Please input 1 to 9 for batch type: "))

# Function to run SAS script
def run_sas(sas_file):
    sas_file_name = os.path.splitext(os.path.basename(sas_file))[0]
    command_string = [sas_path, "-sysin", sas_file, "-config", sas_ver, "-SASUSER", sas_user_path, "-print", f"{sas_file_name}.lst", "-log", f"{sas_file_name}.log"]
    subprocess.run(command_string)

# Loop through rows in Excel sheet
for row in sheet.iter_rows(min_row=3, max_row=3, values_only=True):
    # Extract program name, QC program name, and batch level
    program_name_col = row.index("PROGRAM NAME") + 1
    qc_program_col = row.index("QC PROGRAM") + 1
    batch_level_col = row.index("BATCH LEVEL") + 1

    # Iterate over rows with SAS file paths
    for row in sheet.iter_rows(min_row=4, values_only=True):
        sas_file = row[0]  # Assuming SAS file path is in the first column
        program_name = row[program_name_col - 1]
        qc_program_name = row[qc_program_col - 1]
        batch_level = row[batch_level_col - 1]

        # Run SAS script based on batch level and batch type
        if comb_type == 1 or comb_type == 3 or comb_type == 5 or comb_type == 7 or comb_type == 9:
            lookup_result = program_name
        elif comb_type == 2 or comb_type == 4 or comb_type == 6 or comb_type == 8 or comb_type == 9:
            lookup_result = qc_program_name

        if lookup_result != 0:
            run_sas(sas_file)
