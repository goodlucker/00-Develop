import os
import subprocess
from openpyxl import load_workbook
from tkinter import Tk, filedialog

# 定义是否单独执行每个 SAS 文件
single_batch = False

# SAS 安装路径和版本配置文件路径
sas_path = r"C:\Program Files\SASHome\SASFoundation\9.4\sas.exe"
sas_ver = r"C:\Program Files\SASHome\SASFoundation\9.4\nls\u8\sasv9.cfg"
sas_user_path = r"C:\Program Files\SASHome\SASFoundation\9.4"

# 获取当前脚本所在目录
current_path = os.path.dirname(os.path.abspath(__file__))

# 选择 Tracker 文件
root = Tk()
root.withdraw()  # 隐藏 Tkinter 窗口
track_name = filedialog.askopenfilename(title="Please select the tracker.", filetypes=[("Excel Files", "*.xlsx")])

# 选择 SAS 文件
sas_files = filedialog.askopenfilenames(title="Please select the SAS files.", filetypes=[("SAS Files", "*.sas")])

# 打开 Tracker 文件
wb = load_workbook(filename=track_name)
ws = wb["TLFS"]

# 获取最大层级
max_level = max(ws["R"])

# 遍历层级并运行 SAS 程序
for i in range(1, max_level + 5):
    for sas_file in sas_files:
        file_name = os.path.basename(sas_file)
        level = 0
        for row in ws.iter_rows(values_only=True):
            if row[6] == file_name:
                level = row[11]
                break

        if level == i or (i > max_level and (file_name.upper() == "COMPRPT.SAS" or file_name.upper() == "LOGCHK.SAS")):
            if single_batch:
                subprocess.run([sas_path, "-sysin", sas_file, "-config", sas_ver, "-SASUSER", sas_user_path])
            else:
                subprocess.Popen([sas_path, "-sysin", sas_file, "-config", sas_ver, "-SASUSER", sas_user_path])

# 关闭 Excel 文件
wb.close()

print(f"{len(sas_files)} files complete.")
