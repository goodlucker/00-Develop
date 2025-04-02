import os
from PyPDF2 import PdfReader
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

def extract_bookmarks(reader):
    """Extracts bookmarks from the PDF, returning a dictionary mapping page numbers to bookmark titles."""
    bookmarks = {}

    def _parse_bookmark(outline, parent_title=''):
        if isinstance(outline, list):
            for item in outline:
                _parse_bookmark(item, parent_title)
        else:
            title = f"{parent_title} > {outline.title}" if parent_title else outline.title
            try:
                page_num = reader.get_destination_page_number(outline)
                bookmarks[page_num] = title
            except Exception as e:
                print(f"Error parsing bookmark: {e}")

    try:
        outlines = reader.outline  # Use outline attribute to get bookmarks
        if outlines:
            _parse_bookmark(outlines)
    except Exception as e:
        print(f"Error extracting bookmarks: {e}")

    return bookmarks

def extract_annotations(pdf_path):
    # Open the PDF file
    pypdf_doc = PdfReader(open(pdf_path, "rb"))
    annotations = []

    # Get bookmarks
    bookmarks = extract_bookmarks(pypdf_doc)
    sorted_bookmark_pages = sorted(bookmarks.keys())  # Sort bookmarks by page number

    # Iterate through all pages
    last_bookmark_name = "No Bookmark"
    bookmark_index = 0
    
    for page_num in range(len(pypdf_doc.pages)):
        pypdf_page = pypdf_doc.pages[page_num]
        
        # Update last bookmark if we've passed a bookmark page
        if bookmark_index < len(sorted_bookmark_pages) and page_num >= sorted_bookmark_pages[bookmark_index]:
            last_bookmark_name = bookmarks[sorted_bookmark_pages[bookmark_index]]
            bookmark_index += 1

        # Check if the page contains annotations
        if '/Annots' in pypdf_page:
            annots = pypdf_page['/Annots']
            if isinstance(annots, list):
                rect_annotations = {}

                for annot in annots:
                    annot_obj = annot.get_object()  # Dereference the indirect object
                    subtype = annot_obj.get('/Subtype')
                    contents = annot_obj.get('/Contents', 'No content')
                    rect = tuple(annot_obj.get('/Rect', []))  # Convert ArrayObject to tuple
                    author = annot_obj.get('/T', 'Unknown author')
                    reply_to = annot_obj.get('/IRT')  # Get reply-to reference

                    # Decode bytes to string if necessary
                    if isinstance(contents, bytes):
                        contents = contents.decode('utf-8')
                    if isinstance(author, bytes):
                        author = author.decode('utf-8')

                    # Store annotations by Rect or IRT to associate comments with replies
                    key = rect if reply_to is None else tuple(reply_to.get_object().get('/Rect', rect))

                    if key not in rect_annotations:
                        rect_annotations[key] = []

                    rect_annotations[key].append({
                        "Page": page_num + 1,
                        "Bookmark": last_bookmark_name,  # Use the last bookmark encountered
                        "Annotation Type": subtype,
                        "Content": contents,
                        "Author": author,
                        "Rect": rect,
                        "Is Reply": reply_to is not None
                    })

                # Separate main comments and replies
                for rect, comments in rect_annotations.items():
                    base_comment = None
                    replies = []

                    # Distinguish between main comments and replies
                    for comment in comments:
                        if comment["Is Reply"]:
                            replies.append(comment)
                        else:
                            base_comment = comment

                    if base_comment:
                        annotation_entry = {
                            "Page": base_comment["Page"],
                            "Bookmark": base_comment["Bookmark"],
                            "Annotation Type": base_comment["Annotation Type"],
                            "Main Comment": base_comment["Content"],
                            "Main Author": base_comment["Author"],
                            "PDF Path": pdf_path  # Store PDF path for creating hyperlinks
                        }

                        # Add replies to separate columns
                        for i, reply in enumerate(replies, start=1):
                            annotation_entry[f"Reply {i}"] = reply["Content"]
                            annotation_entry[f"Reply {i} Author"] = reply["Author"]

                        annotations.append(annotation_entry)

    return annotations

def generate_html_link(pdf_path, page_num, html_path):
    """Generate an HTML file that opens a PDF at a specific page."""
    html_content = f"""
    <html>
    <head>
        <script type="text/javascript">
            function openPdf() {{
                window.open('{pdf_path}#page={page_num}', '_blank');
            }}
        </script>
    </head>
    <body onload="openPdf()"></body>
    </html>
    """
    with open(html_path, 'w') as file:
        file.write(html_content)

def save_annotations_to_excel(annotations, output_path):
    df = pd.DataFrame(annotations)

    # Create a subdirectory for HTML files if it doesn't exist
    html_dir = os.path.join(os.path.dirname(output_path), "html_links")
    if not os.path.exists(html_dir):
        os.makedirs(html_dir)

    # Get base PDF file name for HTML link naming
    base_pdf_file_name = os.path.splitext(os.path.basename(annotations[0]["PDF Path"]))[0]

    for index, row in df.iterrows():
        page_num = row["Page"]
        pdf_path = row["PDF Path"].replace("\\", "/")  # Replace backslashes with forward slashes for URL
        html_file_name = f"{base_pdf_file_name}_link_{index}.html"
        html_file_path = os.path.join(html_dir, html_file_name)

        # Generate HTML file with link to the specific page
        generate_html_link(pdf_path, page_num, html_file_path)

        # Create hyperlink to the HTML file
        df.at[index, "Page"] = f'=HYPERLINK("{html_file_path}", "{page_num}")'

    # Drop the PDF Path column as it is not needed in the final output
    df.drop(columns=["PDF Path"], inplace=True)

    # Save as an Excel file
    df.to_excel(output_path, index=False, engine="openpyxl")

    # Adjust column widths and add autofilter using openpyxl
    workbook = load_workbook(output_path)
    worksheet = workbook.active

    # Set autofilter
    worksheet.auto_filter.ref = worksheet.dimensions

    # Set specific column widths (you can adjust these values according to your needs)
    column_widths = {
        "A": 10,  # Page
        "B": 30,  # Bookmark
        "C": 15,  # Annotation Type
        "D": 50,  # Main Comment
        "E": 20,  # Main Author
        "F": 50,  # Reply 1
        "G": 20,  # Reply 1 Author
        # Add more columns if necessary
    }

    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    # Enable text wrapping for all columns
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            # Style hyperlinks
            if cell.hyperlink:
                cell.font = Font(color="0000FF", underline="single")

    # Save the adjusted excel
    workbook.save(output_path)

def browse_pdf_files():
    files = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
    pdf_files.set(";".join(files))

def browse_output_folder():
    folder = filedialog.askdirectory()
    output_folder.set(folder)

def process_pdfs():
    pdf_paths = pdf_files.get().split(";")
    output_dir = output_folder.get()

    if not pdf_paths[0] or not output_dir:
        messagebox.showwarning("Input Error", "Please select both PDF files and output folder.")
        return

    try:
        for pdf_path in pdf_paths:
            annotations = extract_annotations(pdf_path)
            pdf_file_name = os.path.splitext(os.path.basename(pdf_path))[0]
            output_path = os.path.join(output_dir, f"{pdf_file_name}_annotations.xlsx")
            save_annotations_to_excel(annotations, output_path)
        messagebox.showinfo("Success", f"Annotations have been saved to {output_dir}")
    except Exception as e:
        messagebox.showerror("Processing Error", f"An error occurred: {str(e)}")

# Create the main window
root = tk.Tk()
root.title("PDF Annotation Extractor")

# Variables to store file paths
pdf_files = tk.StringVar()
output_folder = tk.StringVar()

# Create and layout the UI elements
tk.Label(root, text="Select PDF Files:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
tk.Entry(root, textvariable=pdf_files, width=50).grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=browse_pdf_files).grid(row=0, column=2, padx=10, pady=10)

tk.Label(root, text="Select Output Folder:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
tk.Entry(root, textvariable=output_folder, width=50).grid(row=1, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=browse_output_folder).grid(row=1, column=2, padx=10, pady=10)

tk.Button(root, text="Extract Annotations", command=process_pdfs).grid(row=2, columnspan=3, pady=20)

# Run the main loop
root.mainloop()