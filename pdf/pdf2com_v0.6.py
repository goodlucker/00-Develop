# pdf2com.py
# !/usr/bin/env python3
# -*- coding: utf-8 -*-
# Author: Rick.Xu
import os
import json
import webbrowser
from PyPDF2 import PdfReader
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font

CONFIG_FILE = "config.json"

def extract_bookmarks(reader):
    """Extracts bookmarks from the PDF, returning a dictionary mapping page numbers to bookmark titles."""
    bookmarks = {}

    def _parse_bookmark(outline, parent_title=''):
        if isinstance(outline, list):
            for item in outline:
                _parse_bookmark(item, parent_title)
        else:
            title = f"{parent_title} > {outline.title}" if parent_title else outline.title
            try:
                page_num = reader.get_destination_page_number(outline)
                bookmarks[page_num] = title
            except Exception as e:
                print(f"Error parsing bookmark: {e}")

    try:
        outlines = reader.outline  # Use outline attribute to get bookmarks
        if outlines:
            _parse_bookmark(outlines)
    except Exception as e:
        print(f"Error extracting bookmarks: {e}")

    return bookmarks

def extract_annotations(pdf_path):
    # Open the PDF file
    with open(pdf_path, "rb") as f:
        pypdf_doc = PdfReader(f)
        annotations = []

        # Get bookmarks
        bookmarks = extract_bookmarks(pypdf_doc)
        sorted_bookmark_pages = sorted(bookmarks.keys())  # Sort bookmarks by page number

        # Iterate through all pages
        last_bookmark_name = "No Bookmark"
        bookmark_index = 0
        
        for page_num in range(len(pypdf_doc.pages)):
            pypdf_page = pypdf_doc.pages[page_num]
            
            # Update last bookmark if we've passed a bookmark page
            if bookmark_index < len(sorted_bookmark_pages) and page_num >= sorted_bookmark_pages[bookmark_index]:
                last_bookmark_name = bookmarks[sorted_bookmark_pages[bookmark_index]]
                bookmark_index += 1

            # Check if the page contains annotations
            if '/Annots' in pypdf_page:
                annots = pypdf_page['/Annots']
                if isinstance(annots, list):
                    rect_annotations = {}

                    for annot in annots:
                        annot_obj = annot.get_object()  # Dereference the indirect object
                        subtype = annot_obj.get('/Subtype')
                        contents = annot_obj.get('/Contents', 'No content')
                        rect = tuple(annot_obj.get('/Rect', []))  # Convert ArrayObject to tuple
                        author = annot_obj.get('/T', 'Unknown author')
                        reply_to = annot_obj.get('/IRT')  # Get reply-to reference

                        # Decode bytes to string if necessary
                        if isinstance(contents, bytes):
                            contents = contents.decode('utf-8')
                        if isinstance(author, bytes):
                            author = author.decode('utf-8')

                        # Store annotations by Rect or IRT to associate comments with replies
                        key = rect if reply_to is None else tuple(reply_to.get_object().get('/Rect', rect))

                        if key not in rect_annotations:
                            rect_annotations[key] = []

                        rect_annotations[key].append({
                            "Page": page_num + 1,
                            "Bookmark": last_bookmark_name,  # Use the last bookmark encountered
                            "Annotation Type": subtype,
                            "Content": contents,
                            "Author": author,
                            "Rect": rect,
                            "Is Reply": reply_to is not None
                        })

                    # Separate main comments and replies
                    for rect, comments in rect_annotations.items():
                        base_comment = None
                        replies = []

                        # Distinguish between main comments and replies
                        for comment in comments:
                            if comment["Is Reply"]:
                                replies.append(comment)
                            else:
                                base_comment = comment

                        if base_comment:
                            annotation_entry = {
                                "Page": base_comment["Page"],
                                "Bookmark": base_comment["Bookmark"],
                                "Annotation Type": base_comment["Annotation Type"],
                                "Main Comment": base_comment["Content"],
                                "Main Author": base_comment["Author"],
                                "PDF Path": pdf_path  # Store PDF path for creating hyperlinks
                            }

                            # Add replies to separate columns
                            for i, reply in enumerate(replies, start=1):
                                annotation_entry[f"Reply {i}"] = reply["Content"]
                                annotation_entry[f"Reply {i} Author"] = reply["Author"]

                            annotations.append(annotation_entry)

    return annotations

def generate_html_link(pdf_path, page_num, html_path):
    """Generate an HTML file that opens a PDF at a specific page."""
    html_content = f"""
    <html>
    <head>
        <script type="text/javascript">
            function openPdf() {{
                window.open('{pdf_path}#page={page_num}', '_blank');
            }}
        </script>
    </head>
    <body onload="openPdf()"></body>
    </html>
    """
    with open(html_path, 'w') as file:
        file.write(html_content)

def save_annotations_to_excel(annotations, workbook, sheet_name, pdf_path):
    df = pd.DataFrame(annotations)

    # Create a subdirectory for HTML files if it doesn't exist
    html_dir = os.path.join(os.path.dirname(pdf_path), "html_links")
    if not os.path.exists(html_dir):
        os.makedirs(html_dir)

    # Get base PDF file name for HTML link naming
    base_pdf_file_name = os.path.splitext(os.path.basename(pdf_path))[0]

    # Create a new sheet
    worksheet = workbook.create_sheet(title=sheet_name[:31])
    
    for index, row in df.iterrows():
        page_num = row["Page"]
        pdf_path_formatted = row["PDF Path"].replace("\\", "/")  # Replace backslashes with forward slashes for URL
        html_file_name = f"{base_pdf_file_name}_link_{index}.html"
        html_file_path = os.path.join(html_dir, html_file_name)

        # Generate HTML file with link to the specific page
        generate_html_link(pdf_path_formatted, page_num, html_file_path)

        # Create hyperlink to the HTML file
        df.at[index, "Page"] = f'=HYPERLINK("{html_file_path}", "{page_num}")'

    # Drop the PDF Path column as it is not needed in the final output
    df.drop(columns=["PDF Path"], inplace=True)

    # Add data to the specified sheet
    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)
    
    # Apply styles and formats
    for r_idx, row in enumerate(worksheet.iter_rows(), 1):
        for c_idx, cell in enumerate(row, 1):
            # Apply wrap text
            cell.alignment = Alignment(wrap_text=True)
            # Style hyperlinks
            if cell.hyperlink:
                cell.font = Font(color="0000FF", underline="single")
    
    # Set auto-filter and column widths
    worksheet.auto_filter.ref = worksheet.dimensions
    column_widths = {
        "A": 10,  # Page
        "B": 30,  # Bookmark
        "C": 15,  # Annotation Type
        "D": 50,  # Main Comment
        "E": 20,  # Main Author
        "F": 50,  # Reply 1
        "G": 20,  # Reply 1 Author
        # Add more columns if necessary
    }
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

def browse_pdf_files():
    files = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
    pdf_files.set(";".join(files))

def browse_output_excel():
    file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    output_file.set(file)

def process_pdfs():
    pdf_paths = pdf_files.get().split(";")
    output_path = output_file.get()

    if not pdf_paths or not output_path:
        messagebox.showwarning("Input Error", "Please select both PDF files and output Excel file.")
        return

    try:
        workbook = Workbook()
        for pdf_path in pdf_paths:
            annotations = extract_annotations(pdf_path)
            pdf_file_name = os.path.splitext(os.path.basename(pdf_path))[0]
            sheet_name = pdf_file_name[:31]  # Ensure sheet name is not too long (max 31 characters)
            save_annotations_to_excel(annotations, workbook, sheet_name, pdf_path)
        # Remove default sheet created by openpyxl
        if "Sheet" in workbook.sheetnames:
            std = workbook["Sheet"]
            workbook.remove(std)

        workbook.save(output_path)
        messagebox.showinfo("Success", f"Annotations have been saved to {output_path}")
    except Exception as e:
        messagebox.showerror("Processing Error", f"An error occurred: {str(e)}")

def save_config():
    config = {
        "pdf_files": pdf_files.get(),
        "output_file": output_file.get()
    }
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=4)
    messagebox.showinfo("Success", "Configuration has been saved.")

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            config = json.load(f)
            pdf_files.set(config.get("pdf_files", ""))
            output_file.set(config.get("output_file", ""))

def send_feedback():
    webbrowser.open("mailto:your_email@example.com?subject=Feedback")

# Create the main window
root = tk.Tk()
root.title("PDF Annotation Extractor")

# Variables to store file paths
pdf_files = tk.StringVar()
output_file = tk.StringVar()

# Load previous configuration if it exists
load_config()

# Create and layout the UI elements
tk.Label(root, text="Select PDF Files:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
tk.Entry(root, textvariable=pdf_files, width=50).grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=browse_pdf_files).grid(row=0, column=2, padx=10, pady=10)

tk.Label(root, text="Select Output Excel File:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
tk.Entry(root, textvariable=output_file, width=50).grid(row=1, column=1, padx=10, pady=10)
tk.Button(root, text="Browse", command=browse_output_excel).grid(row=1, column=2, padx=10, pady=10)

tk.Button(root, text="Extract Annotations", command=process_pdfs).grid(row=2, columnspan=3, pady=10)
tk.Button(root, text="Save Configuration", command=save_config).grid(row=3, columnspan=3, pady=10)

feedback_button = tk.Button(root, text="Feedback", command=send_feedback)
feedback_button.place(relx=1.0, rely=1.0, anchor="se", x=-10, y=-10)

# Run the main loop
root.mainloop()